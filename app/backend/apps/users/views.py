from urllib.parse import urlencode

from django.conf import settings
from django.shortcuts import redirect
from django.utils.crypto import get_random_string
import csv
import io
import logging

from mozilla_django_oidc.utils import add_state_and_verifier_and_nonce_to_session
from mozilla_django_oidc.views import OIDCAuthenticationRequestView as BaseRequestView
from mozilla_django_oidc.views import OIDCAuthenticationCallbackView as BaseCallback

from rest_framework import filters, generics, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework.views import APIView
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.tokens import RefreshToken

from django.db import models as db_models

from .models import RH_ROLES, Augmentation, Evolution, Formation, Notification, Position, Service, Site, User
from .validators import validate_csv_upload
from .serializers import (
    EvolutionSerializer,
    NotificationSerializer,
    PositionSerializer,
    ServiceSerializer,
    SiteSerializer,
    UserMeSerializer,
    UserSerializer,
)

logger = logging.getLogger(__name__)


def get_subordinate_ids(user_id):
    ids = set()
    children = list(User.objects.filter(manager_id=user_id).values_list("id", flat=True))
    for child_id in children:
        ids.add(child_id)
        ids.update(get_subordinate_ids(child_id))
    return ids

class OIDCAuthenticationRequestView(BaseRequestView):
    def get(self, request):
        redirect_uri = getattr(settings, "OIDC_REDIRECT_URI", None)
        if not redirect_uri:
            from django.urls import reverse
            redirect_uri = request.build_absolute_uri(reverse("oidc_authentication_callback"))

        state = get_random_string(32)
        nonce = get_random_string(32)
        params = {
            "response_type": "code",
            "scope": self.get_settings("OIDC_RP_SCOPES", "openid email"),
            "client_id": self.get_settings("OIDC_RP_CLIENT_ID"),
            "redirect_uri": redirect_uri,
            "state": state,
            "nonce": nonce,
        }
        params.update(self.get_extra_params(request))
        add_state_and_verifier_and_nonce_to_session(request, state, params, None)
        request.session["oidc_login_next"] = "/"

        query = urlencode(params)
        authorization_url = self.get_settings("OIDC_OP_AUTHORIZATION_ENDPOINT")
        return redirect(f"{authorization_url}?{query}")

    def get_extra_params(self, request):
        return self.get_settings("OIDC_AUTH_REQUEST_EXTRA_PARAMS", {})


class OIDCCallbackView(BaseCallback):
    def login_failure(self):
        frontend_url = settings.FRONTEND_URL
        return redirect(f"{frontend_url}/login?error=auth_failed")

    def login_success(self):
        frontend_url = settings.FRONTEND_URL
        refresh = RefreshToken.for_user(self.user)
        query = urlencode({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
        })
        return redirect(f"{frontend_url}/auth/callback?{query}")


class MeView(generics.RetrieveUpdateAPIView):
    serializer_class = UserMeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user

    def perform_update(self, serializer):
        if "icon" in serializer.validated_data and serializer.validated_data["icon"]:
            self.request.user.photo = None
            self.request.user.save(update_fields=["photo"])
        serializer.save()

class ProfileAvatarView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        file = request.FILES.get("avatar")
        if not file:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)
        user.photo = file
        user.icon = ""
        user.save(update_fields=["photo", "icon"])
        serializer = UserMeSerializer(user, context={"request": request})
        return Response(serializer.data)


class UserViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ["first_name", "last_name", "email"]

    def get_queryset(self):
        user = self.request.user
        qs = User.objects.select_related("manager", "site")

        if user.role in RH_ROLES:
            qs = qs.all()
        else:
            ids = get_subordinate_ids(user.id)
            if ids:
                ids.add(user.id)
                qs = qs.filter(id__in=ids)
            else:
                qs = qs.filter(id=user.id)

        show_all_statuts = self.request.query_params.get("show_all_statuts")
        if not show_all_statuts:
            qs = qs.exclude(statut__in=("inactif", "sortie"))

        site = self.request.query_params.get("site")
        manager = self.request.query_params.get("manager")
        search = self.request.query_params.get("search")
        if site:
            qs = qs.filter(site_id=site)
        if manager:
            qs = qs.filter(manager_id=manager)
        if search:
            qs = qs.filter(
                db_models.Q(first_name__icontains=search)
                | db_models.Q(last_name__icontains=search)
                | db_models.Q(email__icontains=search)
            )

        return qs

    @action(detail=True, methods=["get"])
    def evolutions(self, request, pk=None):
        user = self.get_object()
        qs = Evolution.objects.filter(employee=user).order_by("date_effet", "created_at")
        return Response(EvolutionSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def export_history_xlsx(self, request, pk=None):
        """Export consolidé des données historisées d'un salarié sur les 6
        dernières années : entretiens professionnels et d'évaluation
        (contenu complet), formations, évolutions, augmentations."""
        from datetime import timedelta

        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        from apps.interviews.models import Interview
        from apps.interviews.views import sanitize_cell_value

        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        employee = self.get_object()
        six_years_ago = timezone.now() - timedelta(days=365.25 * 6)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def write_header(ws, headers):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def write_row(ws, row_idx, values):
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_cell_value(val))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        def autosize(ws):
            ws.freeze_panes = "A2"
            for col_idx, col in enumerate(ws.columns, 1):
                col_letter = chr(64 + col_idx) if col_idx <= 26 else "A"
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        wb = Workbook()

        INTERVIEW_TYPE_SHEETS = [("professional", "Entretiens PRO"), ("annual", "Entretiens Évaluation")]
        interviews = Interview.objects.filter(
            employee=employee, type__in=[t for t, _ in INTERVIEW_TYPE_SHEETS], created_at__gte=six_years_ago,
        ).order_by("-created_at")
        interviews_by_type = {}
        for iv in interviews:
            interviews_by_type.setdefault(iv.type, []).append(iv)

        first_sheet = True
        for iv_type, sheet_title in INTERVIEW_TYPE_SHEETS:
            type_interviews = interviews_by_type.get(iv_type, [])
            ws = wb.active if first_sheet else wb.create_sheet()
            ws.title = sheet_title
            first_sheet = False

            questions_in_type = []
            seen_qids = set()
            for iv in type_interviews:
                for section in iv.content.get("sections", []):
                    for q in section.get("questions", []):
                        qid = q.get("id", "")
                        if qid and qid not in seen_qids:
                            seen_qids.add(qid)
                            questions_in_type.append((qid, q.get("label", qid)))

            headers = ["Statut", "Date limite"] + [label for _qid, label in questions_in_type]
            write_header(ws, headers)

            for row_idx, iv in enumerate(type_interviews, 2):
                answers = {}
                for section in iv.content.get("sections", []):
                    for q in section.get("questions", []):
                        qid = q.get("id", "")
                        answer = q.get("answer")
                        if not qid:
                            continue
                        if q.get("type") == "rating":
                            answers[qid] = str(answer) if answer is not None else ""
                        elif q.get("type") == "table" and isinstance(answer, list):
                            answers[qid] = "; ".join(
                                " | ".join(str(c) for c in row) for row in answer if row
                            )
                        else:
                            answers[qid] = str(answer) if answer else ""

                row_data = [iv.get_status_display(), str(iv.due_date)]
                row_data += [answers.get(qid, "") for qid, _label in questions_in_type]
                write_row(ws, row_idx, row_data)

            if type_interviews:
                last_col = len(headers)
                last_row = len(type_interviews) + 1
                col_letter = chr(64 + last_col) if last_col <= 26 else "A"
                ws.auto_filter.ref = f"A1:{col_letter}{last_row}"
            autosize(ws)

        ws = wb.create_sheet(title="Formations")
        write_header(ws, ["Date", "Domaine", "Libellé", "Nature"])
        formations = Formation.objects.filter(
            employee=employee, date_formation__gte=six_years_ago.date()
        ).order_by("-date_formation")
        for row_idx, f in enumerate(formations, 2):
            write_row(ws, row_idx, [str(f.date_formation) if f.date_formation else "", f.domaine, f.libelle, f.nature])
        autosize(ws)

        ws = wb.create_sheet(title="Augmentations")
        write_header(ws, ["Date d'effet", "Montant"])
        augmentations = Augmentation.objects.filter(
            employee=employee, date_effet__gte=six_years_ago.date()
        ).order_by("-date_effet")
        for row_idx, a in enumerate(augmentations, 2):
            write_row(ws, row_idx, [str(a.date_effet) if a.date_effet else "", str(a.montant) if a.montant is not None else ""])
        autosize(ws)

        ws = wb.create_sheet(title="Évolutions")
        write_header(ws, ["Date d'effet", "Type", "Ancienne valeur", "Nouvelle valeur"])
        evolutions = Evolution.objects.filter(
            employee=employee, date_effet__gte=six_years_ago.date()
        ).order_by("-date_effet")
        for row_idx, e in enumerate(evolutions, 2):
            write_row(ws, row_idx, [
                str(e.date_effet) if e.date_effet else "",
                e.get_type_evolution_display(),
                e.ancienne_valeur,
                e.nouvelle_valeur,
            ])
        autosize(ws)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"historique_{employee.last_name}_{employee.first_name}_6ans.xlsx".replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        from datetime import datetime
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None

    @action(detail=False, methods=["post"])
    def import_csv(self, request):
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        created = 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            email = row.get("email", "").strip().lower()
            if not email:
                errors.append(f"Ligne {row_num}: email manquant")
                continue

            site_name = row.get("site", "").strip()
            site = None
            if site_name:
                site, _ = Site.objects.get_or_create(name=site_name)

            service_name = row.get("service", "").strip()
            service = None
            if service_name:
                service, _ = Service.objects.get_or_create(name=service_name)

            position_name = row.get("position", "").strip()
            position = None
            if position_name:
                position, _ = Position.objects.get_or_create(name=position_name)

            manager_email = row.get("manager_email", "").strip().lower()
            manager = None
            if manager_email:
                manager = User.objects.filter(email=manager_email).first()

            role = row.get("role", "employee").strip().lower()
            if role not in ("admin", "rh", "manager", "employee", "stagiaire", "alternant"):
                role = "employee"

            try:
                user, created_flag = User.objects.get_or_create(
                    email=email,
                    defaults={
                        "username": email,
                        "first_name": row.get("first_name", "").strip(),
                        "last_name": row.get("last_name", "").strip(),
                        "role": role,
                        "service": service,
                        "position": position,
                        "site": site,
                        "manager": manager,
                        "matricule": row.get("matricule", "").strip(),
                        "type_contrat": row.get("type_contrat", "").strip(),
                        "statut": row.get("statut", "actif").strip(),
                        "sexe": row.get("sexe", "").strip(),
                        "telephone": row.get("telephone", "").strip(),
                        "coefficient": row.get("coefficient", "").strip(),
                        "salaire_brut": row.get("salaire_brut", "").strip() or None,
                        "forfait_jour": row.get("forfait_jour", "false").strip().lower() == "true",
                        "tickets_restaurant": row.get("tickets_restaurant", "false").strip().lower() == "true",
                        "cadre": row.get("cadre", "false").strip().lower() == "true",
                        "agence_interim": row.get("agence_interim", "").strip(),
                    },
                )
                if created_flag:
                    user.set_unusable_password()
                    user.save()
                    created += 1
            except Exception as e:
                errors.append(f"Ligne {row_num} ({email}): {e}")

        return Response({"created": created, "errors": errors, "total": len(reader)})

    @action(detail=False, methods=["post"])
    def import_formations(self, request):
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        from .models import Formation
        import csv
        import io
        from datetime import datetime

        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        created = 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            matricule = row.get("Matricule", "").strip()
            if not matricule:
                errors.append(f"Ligne {row_num}: Matricule manquant")
                continue

            employee = User.objects.filter(matricule=matricule).first()
            if not employee:
                errors.append(f"Ligne {row_num}: collaborateur matricule {matricule} introuvable")
                continue

            try:
                date_str = row.get("DATE DE FORMATION", "").strip()
                date_formation = None
                if date_str:
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
                        try:
                            date_formation = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                Formation.objects.create(
                    employee=employee,
                    matricule=matricule,
                    domaine=row.get("DOMAINE", "").strip(),
                    libelle=row.get("Libellé formation", "").strip(),
                    date_formation=date_formation,
                    nature=row.get("NATURE DE LA FORMATION", "").strip(),
                )
                created += 1
            except Exception as e:
                errors.append(f"Ligne {row_num}: {e}")

        return Response({"created": created, "errors": errors})

    @action(detail=False, methods=["post"])
    def import_augmentations(self, request):
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        from .models import Augmentation
        import csv
        import io
        from datetime import datetime

        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        created = 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            matricule = row.get("Matricule", "").strip()
            if not matricule:
                errors.append(f"Ligne {row_num}: Matricule manquant")
                continue

            employee = User.objects.filter(matricule=matricule).first()
            if not employee:
                errors.append(f"Ligne {row_num}: collaborateur matricule {matricule} introuvable")
                continue

            try:
                date_str = row.get("Date", "").strip()
                date_effet = None
                if date_str:
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
                        try:
                            date_effet = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                montant_str = row.get("Montant Augmentation", "").strip().replace(",", ".")
                montant = float(montant_str) if montant_str else None

                Augmentation.objects.create(
                    employee=employee,
                    matricule=matricule,
                    date_effet=date_effet,
                    montant=montant,
                )
                created += 1
            except Exception as e:
                errors.append(f"Ligne {row_num}: {e}")

        return Response({"created": created, "errors": errors})

    @action(detail=False, methods=["post"])
    def import_collaborateurs(self, request):
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        created = 0
        updated = 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            matricule = row.get("Matricule", "").strip()
            if not matricule:
                errors.append(f"Ligne {row_num}: Matricule manquant")
                continue

            try:
                first_name = row.get("Prénom", "").strip().capitalize()
                last_name = row.get("Nom", "").strip().upper()
                date_naissance = self._parse_date(row.get("Date de naissance", "").strip())
                date_entree = self._parse_date(row.get("Date d'entrée", "").strip())

                statut = row.get("Statut", "").strip().lower()
                if statut not in ("actif", "inactif", "sortie"):
                    statut = "actif"

                niveau = row.get("Niveau", "").strip()
                coefficient = row.get("Coefficient", "").strip()

                position_name = row.get("Poste", "").strip()
                position = None
                if position_name:
                    position, _ = Position.objects.get_or_create(name=position_name)

                fonctionnement = row.get("Fonctionnement", "").strip()

                user, created_flag = User.objects.get_or_create(
                    matricule=matricule,
                    defaults={
                        "username": f"collab_{matricule}",
                        "email": f"{matricule}@collaborateur.isb.fr",
                        "first_name": first_name,
                        "last_name": last_name,
                        "role": "employee",
                        "date_naissance": date_naissance,
                        "hire_date": date_entree,
                        "statut": statut,
                        "niveau": niveau,
                        "coefficient": coefficient,
                        "position": position,
                        "fonctionnement": fonctionnement,
                    },
                )
                if created_flag:
                    logger.warning(
                        "Aucun utilisateur trouvé pour le matricule %s, "
                        "création avec email temporaire (%s)",
                        matricule,
                        user.email,
                    )
                    user.set_unusable_password()
                    user.save()
                    created += 1
                else:
                    changed = False
                    for f, v in [
                        ("first_name", first_name),
                        ("last_name", last_name),
                        ("date_naissance", date_naissance),
                        ("hire_date", date_entree),
                        ("statut", statut),
                        ("niveau", niveau),
                        ("coefficient", coefficient),
                        ("position", position),
                        ("fonctionnement", fonctionnement),
                    ]:
                        if getattr(user, f) != v:
                            setattr(user, f, v)
                            changed = True
                    if changed:
                        user.save()
                        updated += 1

            except Exception as e:
                errors.append(f"Ligne {row_num} (matricule {matricule}): {e}")

        return Response({"created": created, "updated": updated, "errors": errors})

    @action(detail=False, methods=["post"])
    def import_kostango(self, request):
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))

        rows = []
        for row_num, row in enumerate(reader, start=2):
            email = row.get("personne email", "").strip().lower()
            if not email:
                continue
            rows.append((row_num, row, email))

        created = 0
        errors = []
        user_map = {}

        # Premier passage : créer tous les utilisateurs
        for row_num, row, email in rows:
            try:
                # Prénom/Nom : d'abord les colonnes dédiées, sinon "personne nom complet"
                first_name = row.get("Prénom", "").strip()
                last_name = row.get("Nom", "").strip()
                if not first_name and not last_name:
                    full = row.get("personne nom complet", "").strip()
                    parts = full.split(" ", 1)
                    if len(parts) == 2:
                        first_name = parts[0].strip().capitalize()
                        last_name = parts[1].strip().upper()
                    elif len(parts) == 1 and parts[0]:
                        last_name = parts[0].strip().upper()

                # Site — on prend le dernier segment après " > "
                site_full = row.get("Site (nom complet)", "").strip()
                site_name = site_full.split(" > ")[-1] if site_full else ""
                site = None
                if site_name:
                    site, _ = Site.objects.get_or_create(name=site_name)

                position_name = row.get("Poste", "").strip()
                position = None
                if position_name:
                    position, _ = Position.objects.get_or_create(name=position_name)

                sexe_map = {"Homme": "homme", "Femme": "femme", "": ""}
                sexe = sexe_map.get(row.get("Sexe", "").strip(), "")

                date_naissance = self._parse_date(row.get("Date de naissance", "").strip())
                hire_date = self._parse_date(row.get("Date d'embauche", "").strip())
                date_sortie = self._parse_date(row.get("Date de sortie", "").strip())

                type_contrat_map = {
                    "CDI": "cdi", "CDD": "cdd",
                    "Intérim": "interim", "INTERIM": "interim",
                    "Alternance": "alternance",
                    "Stage": "stage",
                }
                type_contrat_val = row.get("Type contrat", "").strip()
                type_contrat = type_contrat_map.get(type_contrat_val, "")

                # Statut actif/inactif/sortie
                statut = "actif"
                if date_sortie:
                    sortie_date = self._parse_date(row.get("Date de sortie", "").strip())
                    if sortie_date:
                        statut = "sortie"

                coefficient = row.get("Coefficient", "").strip()

                forfait_jour = row.get("Forfait jour", "false").strip().lower() == "true"
                tickets_restaurant = row.get("Tickets restaurant", "false").strip().lower() == "true"

                agence_interim = row.get("Agence d'intérim", "").strip()

                # Déterminer le rôle
                role = "employee"

                # Déterminer cadre + forfait_jour depuis le statut Kostango
                statut_kostango = row.get("Statut", "").strip()
                cadre = "Cadre" in statut_kostango
                if "FJ" in statut_kostango or "fj" in statut_kostango:
                    forfait_jour = True

                user, created_flag = User.objects.get_or_create(
                    email=email,
                    defaults={
                        "username": email,
                        "first_name": first_name,
                        "last_name": last_name,
                        "role": role,
                        "sexe": sexe,
                        "date_naissance": date_naissance,
                        "site": site,
                        "position": position,
                        "matricule": row.get("Matricule", "").strip(),
                        "type_contrat": type_contrat,
                        "statut": statut,
                        "coefficient": coefficient,
                        "forfait_jour": forfait_jour,
                        "tickets_restaurant": tickets_restaurant,
                        "cadre": cadre,
                        "agence_interim": agence_interim,
                        "hire_date": hire_date,
                        "date_sortie": date_sortie,
                    },
                )
                if created_flag:
                    user.set_unusable_password()
                    user.save()
                    created += 1
                else:
                    # Mise à jour des champs (synchronisation Kostango)
                    changed = False
                    for f, v in [
                        ("first_name", first_name),
                        ("last_name", last_name),
                        ("sexe", sexe),
                        ("date_naissance", date_naissance),
                        ("site", site),
                        ("position", position),
                        ("matricule", row.get("Matricule", "").strip()),
                        ("type_contrat", type_contrat),
                        ("statut", statut),
                        ("coefficient", coefficient),
                        ("forfait_jour", forfait_jour),
                        ("tickets_restaurant", tickets_restaurant),
                        ("cadre", cadre),
                        ("agence_interim", agence_interim),
                        ("hire_date", hire_date),
                        ("date_sortie", date_sortie),
                    ]:
                        if getattr(user, f) != v:
                            setattr(user, f, v)
                            changed = True
                    if changed:
                        user.save()
                user_map[email] = user
            except Exception as e:
                errors.append(f"Ligne {row_num} ({email}): {e}")

        # Construire un mapping nom → email pour chercher les managers par nom
        name_to_email = {}
        for u_email, u in user_map.items():
            full = f"{u.first_name} {u.last_name}".strip().lower()
            if full:
                name_to_email[full] = u_email
            name_to_email[u.email.lower()] = u_email

        # Deuxième passage : assigner les managers via valideur N+1
        for row_num, row, email in rows:
            try:
                user = user_map.get(email)
                if not user:
                    continue
                valideur = row.get("valideur N+1", "").strip().lower()
                if not valideur:
                    continue
                manager_email = name_to_email.get(valideur)
                if manager_email and manager_email in user_map:
                    user.manager = user_map[manager_email]
                    user.save(update_fields=["manager"])
            except Exception as e:
                errors.append(f"Ligne {row_num} ({email}) - manager: {e}")

        # Troisième passage : promouvoir en manager ceux qui sont valideur N+1 de quelqu'un
        manager_emails = set()
        for row_num, row, email in rows:
            valideur = row.get("valideur N+1", "").strip().lower()
            if valideur:
                mgr_email = name_to_email.get(valideur)
                if mgr_email and mgr_email in user_map:
                    manager_emails.add(mgr_email)
        for mgr_email in manager_emails:
            mgr = user_map.get(mgr_email)
            if mgr and mgr.role == "employee":
                mgr.role = "manager"
                mgr.save(update_fields=["role"])

        return Response({"created": created, "errors": errors, "total": len(rows)})

    @action(detail=False, methods=["get"])
    def next_matricule(self, request):
        import re
        existing = User.objects.exclude(matricule="").values_list("matricule", flat=True)
        used = {m for m in existing if re.match(r"^\d+$", m)}
        next_num = 1
        if used:
            next_num = max(int(m) for m in used) + 1
        return Response({"matricule": f"{next_num:08d}"})

    def destroy(self, request, *args, **kwargs):
        return Response(
            {"error": "La suppression d'un utilisateur n'est pas autorisée"},
            status=status.HTTP_403_FORBIDDEN,
        )


class SiteViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Site.objects.all()
    serializer_class = SiteSerializer
    permission_classes = [permissions.IsAuthenticated]


class ServiceViewSet(viewsets.ModelViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def perform_create(self, serializer):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent créer un service")
        serializer.save()

    def perform_update(self, serializer):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent modifier un service")
        serializer.save()

    def perform_destroy(self, instance):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent supprimer un service")
        instance.delete()


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def perform_create(self, serializer):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent créer un poste")
        serializer.save()

    def perform_update(self, serializer):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent modifier un poste")
        serializer.save()

    def perform_destroy(self, instance):
        if self.request.user.role not in RH_ROLES:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seuls les RH/Admin peuvent supprimer un poste")
        instance.delete()


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role in ("admin", "rh"):
            return Notification.objects.none()
        from datetime import date, timedelta
        from apps.interviews.models import Interview
        today = date.today()
        week_end = today + timedelta(days=7)
        for iv in Interview.objects.filter(
            manager=self.request.user,
            status__in=("draft", "in_progress"),
            due_date__gte=today,
            due_date__lte=week_end,
        ):
            Notification.objects.get_or_create(
                user=self.request.user,
                message=f"Échéance dans {(iv.due_date - today).days} jour(s) : {iv.get_type_display()} pour {iv.employee.get_full_name() or iv.employee.email}",
                link=f"/interviews/{iv.id}",
                is_read=False,
            )
        return Notification.objects.filter(user=self.request.user)

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save(update_fields=["is_read"])
        return Response({"status": "ok"})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        self.get_queryset().filter(is_read=False).update(is_read=True)
        return Response({"status": "ok"})


class SSOLoginView(APIView):
    """
    Valide un token SSO émis par ISBibliotheque.
    POST /api/auth/sso/  { "token": "<uuid>" }
    Appelle ISBibliotheque POST /api/sso/consume, récupère l'utilisateur,
    crée/met à jour le compte local, retourne les tokens JWT.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        import requests as http_requests
        from django.contrib.auth import get_user_model

        UserModel = get_user_model()

        token = request.data.get("token", "").strip()
        if not token:
            return Response({"error": "Token manquant"}, status=status.HTTP_400_BAD_REQUEST)

        biblio_url = getattr(settings, "ISBIBLIOTHEQUE_URL", "").rstrip("/")
        if not biblio_url:
            return Response({"error": "ISBIBLIOTHEQUE_URL non configurée"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            resp = http_requests.post(
                f"{biblio_url}/api/sso/consume",
                json={"token": token},
                timeout=10,
            )
        except http_requests.exceptions.RequestException as e:
            return Response({"error": f"Impossible de joindre ISBibliotheque : {e}"}, status=status.HTTP_502_BAD_GATEWAY)

        import logging
        logger = logging.getLogger(__name__)
        logger.warning("SSO consume response: status=%s body=%s", resp.status_code, resp.text[:500])

        if resp.status_code in (401, 404, 410):
            return Response({"error": "Token SSO invalide ou expiré"}, status=status.HTTP_401_UNAUTHORIZED)
        if not resp.ok:
            return Response({"error": f"Erreur ISBibliotheque (HTTP {resp.status_code}): {resp.text[:200]}"}, status=status.HTTP_502_BAD_GATEWAY)

        data = resp.json()
        user_data = data.get("user") or data
        email = (user_data.get("email") or "").strip().lower()
        name = user_data.get("name") or ""
        roles = user_data.get("roles") or []
        is_admin = user_data.get("isAdmin") or False

        if not email:
            return Response({"error": "Email manquant dans la réponse SSO"}, status=status.HTTP_502_BAD_GATEWAY)

        # Déduire le rôle interne
        role_map = {"rh": "rh", "admin": "admin", "manager": "manager"}
        internal_role = "employee"
        if is_admin:
            internal_role = "admin"
        else:
            for r in roles:
                if r in role_map:
                    internal_role = role_map[r]
                    break

        # Prénom / nom
        parts = name.split(" ", 1)
        first_name = parts[0] if parts else ""
        last_name = parts[1] if len(parts) > 1 else ""

        user, created = UserModel.objects.get_or_create(
            email=email,
            defaults={
                "username": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": internal_role,
                "is_active": True,
            },
        )
        if not created:
            # Mettre à jour les infos si l'utilisateur existe déjà
            updated = False
            if first_name and user.first_name != first_name:
                user.first_name = first_name
                updated = True
            if last_name and user.last_name != last_name:
                user.last_name = last_name
                updated = True
            if updated:
                user.save(update_fields=["first_name", "last_name"])

        refresh = RefreshToken.for_user(user)
        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": UserMeSerializer(user).data,
        })


class DevLoginThrottle(AnonRateThrottle):
    scope = "dev_login"


class DevLoginView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [DevLoginThrottle]

    def post(self, request):
        if not settings.DEBUG:
            return Response(status=status.HTTP_404_NOT_FOUND)

        from django.contrib.auth import authenticate

        email = request.data.get("email", "").strip().lower()
        password = request.data.get("password", "")

        user = authenticate(request, username=email, password=password)
        if not user:
            return Response({"error": "Identifiants invalides"}, status=status.HTTP_401_UNAUTHORIZED)

        refresh = RefreshToken.for_user(user)
        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": UserMeSerializer(user).data,
        })


class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        refresh_token = request.data.get("refresh")
        if refresh_token:
            try:
                RefreshToken(refresh_token).blacklist()
            except TokenError:
                # Token deja invalide/expire/blackliste : rien de plus a faire.
                pass
        return Response(status=status.HTTP_204_NO_CONTENT)
