import copy
import csv
import datetime
import io

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.interviews.models import Campaign, Interview, InterviewTemplate
from apps.interviews.templates import ANNUAL_TEMPLATE
from apps.users.models import Site, User


SECTIONS_V1 = [
    {"id": "s1", "title": "Section 1", "questions": [{"id": "q1", "label": "Question 1", "type": "textarea", "answer": ""}]}
]
SECTIONS_V2 = [
    {"id": "s1", "title": "Section 1 modifiée", "questions": [{"id": "q1", "label": "Question 1 modifiée", "type": "textarea", "answer": ""}]}
]


class TemplateSnapshotTests(TestCase):
    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234", role="employee"
        )
        self.template = InterviewTemplate.objects.create(
            name="Template annuel", type="annual", sections=SECTIONS_V1
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_creating_interview_captures_template_snapshot(self):
        response = self.client.post(
            "/api/interviews/",
            {
                "employee": self.employee.id,
                "template": self.template.id,
                "type": "annual",
                "due_date": "2026-12-31",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        interview = Interview.objects.get(pk=response.data["id"])
        self.assertEqual(interview.template_snapshot, SECTIONS_V1)

    def test_modifying_template_after_creation_does_not_change_existing_snapshot(self):
        interview = Interview.objects.create(
            employee=self.employee,
            manager=self.rh_user,
            template=self.template,
            template_snapshot=list(self.template.sections),
            type="annual",
            due_date=datetime.date(2026, 12, 31),
        )

        self.template.sections = SECTIONS_V2
        self.template.save()

        interview.refresh_from_db()
        self.assertEqual(interview.template_snapshot, SECTIONS_V1)
        self.assertEqual(self.template.sections, SECTIONS_V2)

    def test_template_version_increments_on_sections_change(self):
        self.assertEqual(self.template.version, 1)

        self.template.sections = SECTIONS_V2
        self.template.save()
        self.template.refresh_from_db()
        self.assertEqual(self.template.version, 2)

        # Saving again without changing sections must not bump the version.
        self.template.name = "Template annuel (renommé)"
        self.template.save()
        self.template.refresh_from_db()
        self.assertEqual(self.template.version, 2)

    def test_template_version_increments_via_api_patch(self):
        response = self.client.patch(
            f"/api/interview-templates/{self.template.id}/",
            {"sections": SECTIONS_V2},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.template.refresh_from_db()
        self.assertEqual(self.template.version, 2)

    def test_effective_template_sections_falls_back_when_snapshot_missing(self):
        interview = Interview.objects.create(
            employee=self.employee,
            manager=self.rh_user,
            template=self.template,
            type="annual",
            due_date=datetime.date(2026, 12, 31),
        )
        self.assertIsNone(interview.template_snapshot)
        self.assertEqual(interview.get_effective_template_sections(), self.template.sections)

    def test_campaign_generate_captures_template_snapshot(self):
        campaign = Campaign.objects.create(
            name="Campagne annuelle",
            template=self.template,
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
            population_filter={"employees": [self.employee.id]},
        )
        response = self.client.post(f"/api/campaigns/{campaign.id}/generate/")
        self.assertEqual(response.status_code, 200, response.data)

        interview = Interview.objects.get(campaign=campaign, employee=self.employee)
        self.assertEqual(interview.template_snapshot, SECTIONS_V1)


class CampaignPopulationFilterValidationTests(TestCase):
    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234", role="employee"
        )
        self.site = Site.objects.create(name="Paris")
        self.template = InterviewTemplate.objects.create(
            name="Template annuel", type="annual", sections=SECTIONS_V1
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def _campaign_payload(self, population_filter):
        return {
            "name": "Campagne test",
            "template": self.template.id,
            "start_date": "2026-01-01",
            "due_date": "2026-12-31",
            "population_filter": population_filter,
        }

    def test_create_campaign_with_unknown_site_id_rejected(self):
        response = self.client.post(
            "/api/campaigns/", self._campaign_payload({"site": 999999}), format="json"
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("population_filter", response.data)

    def test_create_campaign_with_unknown_employee_id_rejected(self):
        response = self.client.post(
            "/api/campaigns/", self._campaign_payload({"employees": [999999]}), format="json"
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("population_filter", response.data)

    def test_create_campaign_with_valid_filter_accepted(self):
        response = self.client.post(
            "/api/campaigns/",
            self._campaign_payload({"site": self.site.id, "employees": [self.employee.id]}),
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)

    def test_update_campaign_with_unknown_site_id_rejected(self):
        campaign = Campaign.objects.create(
            name="Campagne",
            template=self.template,
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
            population_filter={},
        )
        response = self.client.patch(
            f"/api/campaigns/{campaign.id}/",
            {"population_filter": {"service": 999999}},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("population_filter", response.data)

    def test_generate_with_no_matching_employees_returns_explicit_error(self):
        other_site = Site.objects.create(name="Lyon")
        campaign = Campaign.objects.create(
            name="Campagne vide",
            template=self.template,
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
            population_filter={"site": other_site.id},
        )
        response = self.client.post(f"/api/campaigns/{campaign.id}/generate/")
        self.assertEqual(response.status_code, 400)
        self.assertIn("Aucun collaborateur", response.data.get("error", ""))
        self.assertEqual(Interview.objects.filter(campaign=campaign).count(), 0)


class InterviewManagerDeletionTests(TestCase):
    """Supprimer le manager d'un entretien ne doit plus supprimer
    l'entretien en cascade : le manager doit passer a null."""

    def test_deleting_manager_sets_interview_manager_to_null_instead_of_deleting_it(self):
        manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=manager,
        )
        interview = Interview.objects.create(
            employee=employee, manager=manager,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )

        manager.delete()

        interview.refresh_from_db()
        self.assertIsNone(interview.manager_id)
        self.assertTrue(Interview.objects.filter(pk=interview.pk).exists())


class PreviousYearObjectivesBilanTests(TestCase):
    """A la creation d'un nouvel entretien annuel, les objectifs et
    competences fixes lors de l'entretien annuel precedent (complete/signe)
    doivent etre repris dans des sections de bilan a evaluer."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=self.rh_user,
        )
        self.template = InterviewTemplate.objects.create(
            name="Annuel", type="annual", sections=ANNUAL_TEMPLATE["sections"],
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_new_annual_interview_carries_forward_previous_objectives(self):
        previous = Interview.objects.create(
            employee=self.employee, manager=self.rh_user, type="annual",
            status="completed", due_date=datetime.date(2025, 12, 31),
            template=self.template, content={"sections": copy.deepcopy(ANNUAL_TEMPLATE["sections"])},
        )
        content = previous.content
        for section in content["sections"]:
            if section["id"] == "objectifs":
                for q in section["questions"]:
                    q["answer"] = f"objectif rempli {q['id']}"
        previous.content = content
        previous.save()

        response = self.client.post(
            "/api/interviews/", {
                "employee": self.employee.id, "type": "annual",
                "due_date": "2026-12-31", "template": self.template.id,
            }, format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        new_interview = Interview.objects.get(pk=response.data["id"])
        sections_by_id = {s["id"]: s for s in new_interview.content["sections"]}
        self.assertIn("bilan_objectifs_precedents", sections_by_id)
        bilan_questions = sections_by_id["bilan_objectifs_precedents"]["questions"]
        self.assertTrue(all(q["type"] == "objectif_bilan" for q in bilan_questions))
        self.assertTrue(all(q["objectif_texte"].startswith("objectif rempli") for q in bilan_questions))
        self.assertNotIn("bilan_competences_precedentes", sections_by_id)

    def test_first_annual_interview_has_no_bilan_section(self):
        response = self.client.post(
            "/api/interviews/", {
                "employee": self.employee.id, "type": "annual",
                "due_date": "2026-12-31", "template": self.template.id,
            }, format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        interview = Interview.objects.get(pk=response.data["id"])
        section_ids = [s["id"] for s in interview.content["sections"]]
        self.assertNotIn("bilan_objectifs_precedents", section_ids)


def _find_question(sections, question_id):
    for section in sections:
        for question in section.get("questions", []):
            if question.get("id") == question_id:
                return question
    return None


class ObjectifTablesTests(TestCase):
    """L'entretien d'évaluation contient un tableau "Objectif à évaluer"
    (pré-rempli avec les objectifs définis l'année précédente) et un
    tableau "Objectif à définir" (vide)."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=self.manager,
        )
        self.template = InterviewTemplate.objects.create(
            name="Annuel", type="annual", sections=copy.deepcopy(ANNUAL_TEMPLATE["sections"]),
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_annual_template_has_both_objectif_tables(self):
        evaluer = _find_question(ANNUAL_TEMPLATE["sections"], "objectif_a_evaluer")
        definir = _find_question(ANNUAL_TEMPLATE["sections"], "objectif_a_definir")
        expected_col_ids = [
            "theme", "objectif", "date_realisation", "niveau_realisation",
            "note_collaborateur", "remarque_collaborateur", "note_manager", "remarque_manager",
        ]
        for question in (evaluer, definir):
            self.assertIsNotNone(question)
            self.assertEqual(question["type"], "table")
            self.assertEqual([c["id"] for c in question["columns"]], expected_col_ids)

    def test_first_annual_interview_has_empty_objectif_tables(self):
        response = self.client.post(
            "/api/interviews/", {
                "employee": self.employee.id, "type": "annual",
                "due_date": "2026-12-31", "template": self.template.id,
            }, format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        interview = Interview.objects.get(pk=response.data["id"])
        evaluer = _find_question(interview.content["sections"], "objectif_a_evaluer")
        definir = _find_question(interview.content["sections"], "objectif_a_definir")
        self.assertTrue(all(cell is None for row in evaluer["answer"] for cell in row))
        self.assertTrue(all(cell is None for row in definir["answer"] for cell in row))

    def test_new_annual_interview_prefills_objectif_a_evaluer_from_previous_definir(self):
        previous_rows = [
            ["Qualité", "Réduire les défauts", "", "", None, "", None, ""],
            ["Formation", "Suivre une certification", "", "", None, "", None, ""],
        ]
        previous = Interview.objects.create(
            employee=self.employee, manager=self.manager, type="annual",
            status="completed", due_date=datetime.date(2025, 12, 31),
            template=self.template, content={"sections": copy.deepcopy(ANNUAL_TEMPLATE["sections"])},
        )
        content = previous.content
        definir = _find_question(content["sections"], "objectif_a_definir")
        definir["answer"] = previous_rows
        previous.content = content
        previous.save()

        response = self.client.post(
            "/api/interviews/", {
                "employee": self.employee.id, "type": "annual",
                "due_date": "2026-12-31", "template": self.template.id,
            }, format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        new_interview = Interview.objects.get(pk=response.data["id"])
        evaluer = _find_question(new_interview.content["sections"], "objectif_a_evaluer")
        self.assertEqual(evaluer["answer"], previous_rows)

    def test_campaign_generate_prefills_objectif_a_evaluer_from_previous_definir(self):
        previous_rows = [["Qualité", "Réduire les défauts", "", "", None, "", None, ""]]
        previous = Interview.objects.create(
            employee=self.employee, manager=self.manager, type="annual",
            status="signed", due_date=datetime.date(2025, 12, 31),
            template=self.template, content={"sections": copy.deepcopy(ANNUAL_TEMPLATE["sections"])},
        )
        content = previous.content
        definir = _find_question(content["sections"], "objectif_a_definir")
        definir["answer"] = previous_rows
        previous.content = content
        previous.save()

        campaign = Campaign.objects.create(
            name="Campagne annuelle 2026",
            template=self.template,
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
            population_filter={"employees": [self.employee.id]},
        )
        response = self.client.post(f"/api/campaigns/{campaign.id}/generate/")
        self.assertEqual(response.status_code, 200, response.data)
        interview = Interview.objects.get(campaign=campaign, employee=self.employee)
        evaluer = _find_question(interview.content["sections"], "objectif_a_evaluer")
        self.assertEqual(evaluer["answer"], previous_rows)


class PrintTemplateTests(TestCase):
    """Le titre imprime doit refleter le type reel de l'entretien, un logo
    doit etre present, et les blocs d'historique doivent apparaitre pour un
    entretien de type bilan avec de l'historique disponible."""

    def setUp(self):
        self.manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=self.manager,
        )
        self.client = APIClient()

    def test_print_title_matches_interview_type_for_bilan(self):
        Interview.objects.create(
            employee=self.employee, manager=self.manager, type="professional",
            status="completed", due_date=datetime.date(2026, 1, 1),
            content={"employee_snapshot": {"salaire_brut": "2500"}},
        )
        interview = Interview.objects.create(
            employee=self.employee, manager=self.manager, type="bilan",
            due_date=datetime.date(2026, 12, 31), content={},
        )
        self.client.force_authenticate(user=self.manager)
        response = self.client.get(f"/api/interviews/{interview.id}/print/")
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("ENTRETIEN DE BILAN", html.upper())
        self.assertNotIn("ENTRETIEN D'ÉVALUATION", html.upper())
        self.assertIn('<img class="logo"', html)
        self.assertIn("Parcours professionnel", html)
        self.assertIn("Historique des entretiens", html)


class EmployeeWithoutManagerTests(TestCase):
    """Un salarie classique sans N+1 (manager=None) doit pouvoir avoir des
    entretiens crees, imprimes et generes via campagne sans erreur — le
    manager de l'entretien retombe alors sur le createur (RH)."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=None,
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_create_and_print_interview_without_manager(self):
        response = self.client.post(
            "/api/interviews/", {
                "employee": self.employee.id, "type": "annual",
                "due_date": "2026-12-31", "content": {},
            }, format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        interview = Interview.objects.get(pk=response.data["id"])
        self.assertEqual(interview.manager_id, self.rh_user.id)

        print_response = self.client.get(f"/api/interviews/{interview.id}/print/")
        self.assertEqual(print_response.status_code, 200)

    def test_campaign_generate_falls_back_to_creator_manager(self):
        template = InterviewTemplate.objects.create(
            name="T", type="annual", sections=[{"id": "s1", "title": "S", "questions": []}],
        )
        campaign = Campaign.objects.create(
            name="Camp", template=template,
            start_date=datetime.date(2026, 1, 1), due_date=datetime.date(2026, 12, 31),
            population_filter={"employees": [self.employee.id]},
        )
        response = self.client.post(f"/api/campaigns/{campaign.id}/generate/")
        self.assertEqual(response.status_code, 200, response.data)
        interview = Interview.objects.get(campaign=campaign, employee=self.employee)
        self.assertEqual(interview.manager_id, self.rh_user.id)


class InterviewManagerReassignmentTests(TestCase):
    """Changer le manager (N+1) d'un employe doit resynchroniser
    automatiquement le champ Interview.manager de ses entretiens, pour que
    l'ancien manager perde l'acces et que le nouveau l'obtienne sans action
    manuelle (cf. reassign_managers, qui reste un rattrapage explicite)."""

    def test_changing_employee_manager_reassigns_existing_interviews(self):
        old_manager = User.objects.create_user(
            username="mgr_old", email="mgr_old@example.com", password="pass1234", role="manager"
        )
        new_manager = User.objects.create_user(
            username="mgr_new", email="mgr_new@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=old_manager,
        )
        interview = Interview.objects.create(
            employee=employee, manager=old_manager,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )

        employee.manager = new_manager
        employee.save()

        interview.refresh_from_db()
        self.assertEqual(interview.manager_id, new_manager.id)

    def test_removing_employee_manager_clears_interview_manager(self):
        manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=manager,
        )
        interview = Interview.objects.create(
            employee=employee, manager=manager,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )

        employee.manager = None
        employee.save()

        interview.refresh_from_db()
        self.assertIsNone(interview.manager_id)

    def test_unrelated_user_field_change_does_not_touch_interview_manager(self):
        manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=manager,
        )
        interview = Interview.objects.create(
            employee=employee, manager=manager,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )

        employee.telephone = "0600000000"
        employee.save()

        interview.refresh_from_db()
        self.assertEqual(interview.manager_id, manager.id)


class InterviewEmployeeDeletionTests(TestCase):
    """Un employe ayant un historique d'entretiens ne doit pas pouvoir etre
    supprime silencieusement (PROTECT) : la suppression doit echouer
    explicitement plutot que de detruire l'historique en cascade."""

    def test_deleting_employee_with_interview_history_is_blocked(self):
        from django.db.models import ProtectedError

        manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=manager,
        )
        interview = Interview.objects.create(
            employee=employee, manager=manager,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )

        with self.assertRaises(ProtectedError):
            employee.delete()

        self.assertTrue(Interview.objects.filter(pk=interview.pk).exists())
        self.assertTrue(User.objects.filter(pk=employee.pk).exists())


class ExcelExportFormulaInjectionTests(TestCase):
    """Une reponse d'entretien commencant par un caractere declencheur de
    formule (=, +, -, @) ne doit pas etre ecrite telle quelle dans l'export
    Excel : elle doit etre neutralisee (prefixee par une apostrophe)."""

    def test_malicious_answer_is_sanitized_in_contents_export(self):
        rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234", role="manager"
        )
        employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=manager,
        )
        campaign = Campaign.objects.create(
            name="Campagne export",
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
        )
        Interview.objects.create(
            employee=employee, manager=manager, campaign=campaign,
            type="annual", due_date=datetime.date(2026, 12, 31),
            content={
                "sections": [
                    {
                        "id": "s1",
                        "title": "Section 1",
                        "questions": [
                            {
                                "id": "q1",
                                "label": "Commentaire",
                                "type": "textarea",
                                "answer": "=cmd|'/c calc'!A0",
                            }
                        ],
                    }
                ]
            },
        )

        client = APIClient()
        client.force_authenticate(user=rh_user)
        response = client.get(f"/api/campaigns/{campaign.id}/export_contents_xlsx/")
        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        answer_col = None
        for cell in ws[1]:
            if cell.value == "Commentaire":
                answer_col = cell.column
        self.assertIsNotNone(answer_col)

        cell_value = ws.cell(row=2, column=answer_col).value
        self.assertTrue(cell_value.startswith("'="))
        self.assertIn("cmd", cell_value)


class CampaignExportMatriculeTests(TestCase):
    """Les exports de campagne doivent inclure le matricule du collaborateur
    et celui de son manager."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.manager = User.objects.create_user(
            username="mgr1", email="mgr1@example.com", password="pass1234",
            role="manager", matricule="00000700",
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234",
            role="employee", manager=self.manager, matricule="00000701",
        )
        self.campaign = Campaign.objects.create(
            name="Campagne matricules",
            start_date=datetime.date(2026, 1, 1),
            due_date=datetime.date(2026, 12, 31),
        )
        self.interview = Interview.objects.create(
            employee=self.employee, manager=self.manager, campaign=self.campaign,
            type="annual", due_date=datetime.date(2026, 12, 31),
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_export_contents_xlsx_includes_matricules(self):
        response = self.client.get(f"/api/campaigns/{self.campaign.id}/export_contents_xlsx/")
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertIn("Matricule", headers)
        self.assertIn("Matricule manager", headers)
        matricule_col = headers.index("Matricule") + 1
        manager_matricule_col = headers.index("Matricule manager") + 1
        self.assertEqual(ws.cell(row=2, column=matricule_col).value, "00000701")
        self.assertEqual(ws.cell(row=2, column=manager_matricule_col).value, "00000700")

    def test_export_csv_includes_matricules(self):
        response = self.client.get(
            "/api/interviews/export_csv/", {"campaign_id": self.campaign.id}
        )
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(content)))
        header = rows[0]
        self.assertIn("Matricule", header)
        self.assertIn("Matricule manager", header)


class InterviewTemplateSectionsValidationTests(TestCase):
    """InterviewTemplate.sections doit respecter une structure minimale
    (id/title/questions, question id/label/type) plutot que d'accepter
    n'importe quel JSON."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def _payload(self, sections):
        return {"name": "Template test", "type": "annual", "sections": sections}

    def test_valid_sections_accepted(self):
        response = self.client.post("/api/interview-templates/", self._payload(SECTIONS_V1), format="json")
        self.assertEqual(response.status_code, 201, response.data)

    def test_section_missing_id_rejected(self):
        sections = [{"title": "Section sans id", "questions": []}]
        response = self.client.post("/api/interview-templates/", self._payload(sections), format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sections", response.data)

    def test_section_with_non_list_questions_rejected(self):
        sections = [{"id": "s1", "title": "Section", "questions": "pas une liste"}]
        response = self.client.post("/api/interview-templates/", self._payload(sections), format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sections", response.data)

    def test_question_missing_label_rejected(self):
        sections = [{"id": "s1", "title": "Section", "questions": [{"id": "q1"}]}]
        response = self.client.post("/api/interview-templates/", self._payload(sections), format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sections", response.data)

    def test_question_with_invalid_type_rejected(self):
        sections = [
            {
                "id": "s1", "title": "Section",
                "questions": [{"id": "q1", "label": "Q1", "type": "not-a-real-type"}],
            }
        ]
        response = self.client.post("/api/interview-templates/", self._payload(sections), format="json")
        self.assertEqual(response.status_code, 400)
        self.assertIn("sections", response.data)


class DefaultCommentSectionAndTableRowsTests(TestCase):
    """Chaque entretien cree doit recevoir automatiquement une section
    Commentaire (collaborateur + manager), et toute question de type table
    sans reponse doit demarrer avec 5 lignes vides par defaut."""

    def setUp(self):
        self.rh_user = User.objects.create_user(
            username="rh1", email="rh1@example.com", password="pass1234", role="rh"
        )
        self.employee = User.objects.create_user(
            username="emp1", email="emp1@example.com", password="pass1234", role="employee"
        )
        sections = [
            {
                "id": "s1", "title": "Section 1",
                "questions": [
                    {
                        "id": "q1", "label": "Tableau", "type": "table",
                        "columns": [{"id": "c1", "label": "Colonne 1", "type": "textarea"}],
                        "answer": [],
                    },
                ],
            }
        ]
        self.template = InterviewTemplate.objects.create(
            name="Template avec tableau", type="professional", sections=sections
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.rh_user)

    def test_new_interview_has_default_comment_section(self):
        response = self.client.post(
            "/api/interviews/",
            {
                "employee": self.employee.id,
                "template": self.template.id,
                "type": "professional",
                "due_date": "2026-12-31",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        sections = response.data["content"]["sections"]
        comment_section = next((s for s in sections if s["id"] == "commentaires"), None)
        self.assertIsNotNone(comment_section)
        question_ids = {q["id"] for q in comment_section["questions"]}
        self.assertEqual(question_ids, {"commentaire_collaborateur", "commentaire_manager"})

    def test_new_interview_table_question_gets_five_default_empty_rows(self):
        response = self.client.post(
            "/api/interviews/",
            {
                "employee": self.employee.id,
                "template": self.template.id,
                "type": "professional",
                "due_date": "2026-12-31",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        sections = response.data["content"]["sections"]
        table_q = next(q for s in sections for q in s["questions"] if q["id"] == "q1")
        self.assertEqual(len(table_q["answer"]), 5)
        self.assertTrue(all(cell is None for row in table_q["answer"] for cell in row))

    def test_template_sections_not_mutated_by_default_application(self):
        self.client.post(
            "/api/interviews/",
            {
                "employee": self.employee.id,
                "template": self.template.id,
                "type": "professional",
                "due_date": "2026-12-31",
            },
            format="json",
        )
        self.template.refresh_from_db()
        table_q = self.template.sections[0]["questions"][0]
        self.assertEqual(table_q["answer"], [])
