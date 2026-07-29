import copy
from datetime import timedelta
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Campaign, Interview, InterviewTemplate
from .serializers import CampaignSerializer, InterviewSerializer, InterviewTemplateSerializer
from apps.users.models import RH_ROLES, User
from apps.users.validators import validate_csv_upload

FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def sanitize_cell_value(value):
    """Neutralise l'injection de formule Excel : une cellule dont la valeur
    commence par un caractere declencheur de formule est prefixee par une
    apostrophe pour forcer son interpretation comme texte a l'ouverture."""
    if isinstance(value, str) and value.startswith(FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


BILAN_SOURCE_SECTIONS = [
    ("objectifs", "bilan_objectifs_precedents", "Bilan des objectifs de l'année précédente"),
    ("competences", "bilan_competences_precedentes", "Bilan des compétences de l'année précédente"),
]


def build_previous_year_bilan_sections(employee, interview_type):
    """Pour un entretien annuel, reprend les objectifs/competences fixes
    lors de l'entretien annuel precedent (par section id "objectifs" /
    "competences" du template par defaut) et construit des sections de
    bilan a evaluer (atteint/partiellement/non atteint + commentaire)."""
    if interview_type != "annual":
        return []

    previous = (
        Interview.objects.filter(employee=employee, type="annual", status__in=("completed", "signed"))
        .order_by("-created_at")
        .first()
    )
    if not previous:
        return []

    prev_sections_by_id = {s.get("id"): s for s in previous.content.get("sections", [])}

    bilan_sections = []
    for source_id, new_id, new_title in BILAN_SOURCE_SECTIONS:
        source = prev_sections_by_id.get(source_id)
        if not source or not source.get("questions"):
            continue
        questions = [
            {
                "id": f"bilan_{q.get('id')}",
                "label": q.get("label", q.get("id")),
                "type": "objectif_bilan",
                "objectif_texte": q.get("answer") or "",
                "answer": {"statut": "", "commentaire": ""},
            }
            for q in source["questions"]
            if q.get("answer")
        ]
        if questions:
            bilan_sections.append({"id": new_id, "title": new_title, "questions": questions})
    return bilan_sections


DEFAULT_TABLE_ROWS = 5

COMMENT_SECTION_ID = "commentaires"


def apply_default_sections(sections):
    """Complete une liste de sections d'entretien avant creation :
    - toute question de type "table" sans reponse demarre avec
      DEFAULT_TABLE_ROWS lignes vides (au lieu de 0) ;
    - une section "Commentaire" (commentaire collaborateur + manager) est
      ajoutee par defaut si elle n'est pas deja presente."""
    sections = copy.deepcopy(list(sections))
    for section in sections:
        for question in section.get("questions", []):
            if question.get("type") == "table" and not question.get("answer"):
                nb_cols = len(question.get("columns") or [{}])
                question["answer"] = [[None] * nb_cols for _ in range(DEFAULT_TABLE_ROWS)]

    if not any(s.get("id") == COMMENT_SECTION_ID for s in sections):
        sections.append({
            "id": COMMENT_SECTION_ID,
            "title": "Commentaire",
            "questions": [
                {"id": "commentaire_collaborateur", "label": "Commentaire du collaborateur", "type": "textarea", "answer": ""},
                {"id": "commentaire_manager", "label": "Commentaire du manager", "type": "textarea", "answer": ""},
            ],
        })
    return sections


class InterviewPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if view.action == "stats":
            return True
        basename = getattr(view, 'basename', '')
        if basename in ('interviewtemplate', 'campaign'):
            return request.user.role in ("admin", "rh", "manager")
        if view.action in ("create", "destroy"):
            return request.user.role in RH_ROLES
        return True

    def has_object_permission(self, request, view, obj):
        if request.user.role in RH_ROLES:
            return True
        basename = getattr(view, 'basename', '')
        if basename in ('interviewtemplate', 'campaign'):
            return request.user.role in ("admin", "rh", "manager")
        if view.action in ("retrieve", "print", "pdf"):
            if obj.employee == request.user or obj.manager == request.user:
                return True
            from apps.users.views import get_subordinate_ids
            ids = get_subordinate_ids(request.user.id)
            if obj.employee_id in ids:
                return True
            return False
        if view.action in ("update", "partial_update", "upload_document", "remove_document"):
            if request.user.role in RH_ROLES:
                return True
            if obj.manager == request.user:
                return True
            if obj.employee == request.user and not obj.employee.manager:
                return True
            from apps.users.views import get_subordinate_ids
            if obj.employee_id in get_subordinate_ids(request.user.id):
                return True
            return False
        return False


class InterviewViewSet(viewsets.ModelViewSet):
    serializer_class = InterviewSerializer
    permission_classes = [permissions.IsAuthenticated, InterviewPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["employee__first_name", "employee__last_name", "employee__email"]
    ordering_fields = ["due_date", "created_at", "updated_at", "status"]
    ordering = ["-due_date"]

    def get_object(self):
        from django.shortcuts import get_object_or_404
        obj = get_object_or_404(Interview.objects.select_related("employee", "manager", "template", "campaign"), pk=self.kwargs["pk"])
        self.check_object_permissions(self.request, obj)
        return obj

    def get_queryset(self):
        user = self.request.user
        qs = Interview.objects.select_related("employee", "manager", "template", "campaign")

        scope = self.request.query_params.get("scope")

        if user.role in RH_ROLES:
            qs = qs.all()

        elif user.role == "manager":
            if scope == "own":
                qs = qs.filter(employee=user)
            elif scope == "team":
                from apps.users.views import get_subordinate_ids
                ids = get_subordinate_ids(user.id)
                if ids:
                    qs = qs.filter(employee_id__in=ids)
                else:
                    return qs.none()
            else:
                subordinates = User.objects.filter(manager=user).values_list("id", flat=True)
                if subordinates:
                    qs = qs.filter(employee_id__in=subordinates)
                else:
                    qs = qs.filter(employee=user)

        elif user.role == "employee" and not user.manager:
            qs = qs.filter(employee=user)
        else:
            qs = qs.filter(employee=user)

        type_filter = self.request.query_params.get("type")
        status_filter = self.request.query_params.get("status")
        campaign_filter = self.request.query_params.get("campaign")
        if type_filter:
            qs = qs.filter(type=type_filter)
        if status_filter:
            status_list = status_filter.split(",")
            qs = qs.filter(status__in=status_list)
        if campaign_filter:
            qs = qs.filter(campaign_id=campaign_filter)

        return qs

    def perform_create(self, serializer):
        template = serializer.validated_data.get("template")
        employee = serializer.validated_data.get("employee")
        interview_type = serializer.validated_data.get("type")
        content = serializer.validated_data.get("content")
        if not content:
            content = {}
        if template and not content.get("sections"):
            content["sections"] = list(template.sections)
        if template:
            serializer.validated_data["template_snapshot"] = list(template.sections)
        if employee:
            content["employee_snapshot"] = {
                "position": employee.position.name if employee.position else None,
                "service": employee.service.name if employee.service else None,
                "site": employee.site.name if employee.site else None,
                "coefficient": employee.coefficient,
                "salaire_brut": str(employee.salaire_brut) if employee.salaire_brut else None,
            }
        if employee and interview_type:
            bilan_sections = build_previous_year_bilan_sections(employee, interview_type)
            if bilan_sections:
                content["sections"] = bilan_sections + list(content.get("sections", []))
        content["sections"] = apply_default_sections(content.get("sections", []))
        serializer.validated_data["content"] = content
        serializer.save(manager=employee.manager if employee else None)

    def perform_update(self, serializer):
        employee = serializer.instance.employee
        content = serializer.validated_data.get("content", serializer.instance.content or {})
        if isinstance(content, dict):
            content["employee_snapshot"] = {
                "position": employee.position.name if employee.position else None,
                "service": employee.service.name if employee.service else None,
                "site": employee.site.name if employee.site else None,
                "coefficient": employee.coefficient,
                "salaire_brut": str(employee.salaire_brut) if employee.salaire_brut else None,
            }
        serializer.validated_data["content"] = content
        serializer.save()

    @action(detail=False, methods=["get"])
    def stats(self, request):
        qs = self.get_queryset()
        now = timezone.now().date()
        return Response({
            "total": qs.count(),
            "by_status": qs.values("status").annotate(count=Count("id")),
            "by_type": qs.values("type").annotate(count=Count("id")),
            "overdue": qs.filter(status__in=("draft", "in_progress"), due_date__lt=now).count(),
            "upcoming": qs.filter(status__in=("draft", "in_progress"), due_date__gte=now).count(),
        })

    @action(detail=False, methods=["get"])
    def employees(self, request):
        if request.user.role in RH_ROLES:
            users = User.objects.filter(is_active=True).exclude(statut__in=("inactif", "sortie")).values("id", "first_name", "last_name", "email")
            return Response(list(users))
        return Response([])

    @action(detail=False, methods=["post"])
    def import_historique(self, request):
        """Importe en masse des entretiens historiques (métadonnées
        uniquement, sans contenu détaillé) : État, Date prévue,
        Date de réalisation, Matricule. Le type d'entretien est choisi une
        fois pour tout le fichier."""
        if request.user.role not in RH_ROLES:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)

        interview_type = request.data.get("type", "").strip()
        if interview_type not in dict(Interview.Type.choices):
            return Response({"error": "Type d'entretien invalide"}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Fichier CSV requis"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        import csv
        import io

        reader = csv.DictReader(io.StringIO(file.read().decode("utf-8-sig")))

        status_map = {
            "cloture": "signed", "clôture": "signed", "cloturé": "signed", "clôturé": "signed",
            "realise": "completed", "réalisé": "completed", "realisé": "completed", "réalise": "completed",
        }

        created = 0
        errors = []
        for row_num, row in enumerate(reader, start=2):
            matricule = row.get("Matricule", "").strip()
            if not matricule:
                errors.append(f"Ligne {row_num}: matricule manquant")
                continue
            employee = User.objects.filter(matricule=matricule).first()
            if not employee:
                errors.append(f"Ligne {row_num}: aucun collaborateur avec le matricule {matricule}")
                continue

            try:
                date_prevue = self._parse_date(row.get("Date prévue", "").strip())
                date_realisation = self._parse_date(row.get("Date de réalisation", "").strip())
                statut = status_map.get(row.get("État", "").strip().lower(), "completed")

                iv = Interview.objects.create(
                    employee=employee,
                    manager=employee.manager,
                    type=interview_type,
                    status=statut,
                    due_date=date_prevue or date_realisation or timezone.now().date(),
                    date_realisation=date_realisation,
                    content={},
                )
                if date_realisation:
                    Interview.objects.filter(pk=iv.pk).update(created_at=date_realisation)
                created += 1
            except Exception as e:
                errors.append(f"Ligne {row_num} ({matricule}): {e}")

        return Response({"created": created, "errors": errors})

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        from datetime import datetime
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None

    @action(detail=False, methods=["get"])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        qs = self.get_queryset().filter(status__in=("completed", "signed"))
        campaign_id = request.query_params.get("campaign_id")
        if campaign_id:
            qs = qs.filter(campaign_id=campaign_id)
        filename = "entretiens.csv"
        if campaign_id:
            campaign = Campaign.objects.filter(pk=campaign_id).first()
            if campaign:
                filename = f"entretiens_{campaign.name}.csv".replace(" ", "_")
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["ID", "Employé", "Email", "Manager", "Type", "Statut", "Date limite", "Date création", "Poste", "Service", "Site"])
        for iv in qs.select_related("employee", "manager", "employee__position", "employee__service", "employee__site"):
            snap = iv.content.get("employee_snapshot", {})
            writer.writerow([
                iv.id,
                iv.employee.get_full_name() or iv.employee.email,
                iv.employee.email,
                (iv.manager.get_full_name() or iv.manager.email) if iv.manager else "",
                iv.get_type_display(),
                iv.get_status_display(),
                iv.due_date,
                iv.created_at.date(),
                snap.get("position") or (iv.employee.position.name if iv.employee.position else None),
                snap.get("service") or (iv.employee.service.name if iv.employee.service else None),
                snap.get("site") or (iv.employee.site.name if iv.employee.site else None),
            ])
        return response

    @staticmethod
    def _compute_anciennete(hire_date):
        if not hire_date:
            return None
        from datetime import date

        today = date.today()
        years = today.year - hire_date.year
        months = today.month - hire_date.month
        if today.day < hire_date.day:
            months -= 1
        if months < 0:
            months += 12
            years -= 1

        parts = []
        if years > 0:
            parts.append(f"{years} an" + ("s" if years > 1 else ""))
        if months > 0 or not parts:
            parts.append(f"{months} mois")
        return " ".join(parts)

    def _get_print_context(self, interview):
        sections = interview.content.get("sections", [])
        all_past = list(Interview.objects.filter(employee=interview.employee, status__in=("completed", "signed")).exclude(pk=interview.pk).select_related("manager", "template").order_by("-created_at"))

        if interview.type == "bilan":
            career = Interview.objects.filter(
                employee=interview.employee,
                status__in=("completed", "signed"),
            ).exclude(pk=interview.pk).order_by("created_at")
        else:
            career = Interview.objects.filter(
                employee=interview.employee,
                type="professional",
            ).order_by("created_at")

        salary_chrono = []
        training_history = []
        for iv in reversed(all_past):
            snap = iv.content.get("employee_snapshot", {})
            sal = snap.get("salaire_brut")
            if sal:
                salary_chrono.append({"date": iv.created_at, "type": iv.get_type_display(), "salary": sal, "coefficient": snap.get("coefficient", "")})
            entries = []
            for section in iv.content.get("sections", []):
                for q in section.get("questions", []):
                    qid = q.get("id", "")
                    if any(kw in qid for kw in ["formation", "cpf", "vae"]):
                        if q.get("answer"):
                            entries.append({"label": q.get("label", qid), "answer": q.get("answer", "")})
            if entries:
                training_history.append({"date": iv.created_at, "type": iv.get_type_display(), "entries": entries})
        training_history.reverse()

        return {
            "interview": interview,
            "sections": sections,
            "template_sections": interview.get_effective_template_sections(),
            "history": all_past[:6],
            "career": career,
            "training_history": training_history,
            "salary_history": salary_chrono,
            "logo_data_uri": self._get_logo_data_uri(),
            "anciennete": self._compute_anciennete(interview.employee.hire_date),
        }

    _logo_data_uri_cache = None

    @classmethod
    def _get_logo_data_uri(cls):
        if cls._logo_data_uri_cache is None:
            import base64
            import os

            logo_path = os.path.join(os.path.dirname(__file__), "static", "interviews", "logo.png")
            with open(logo_path, "rb") as f:
                cls._logo_data_uri_cache = "data:image/png;base64," + base64.b64encode(f.read()).decode("ascii")
        return cls._logo_data_uri_cache

    @action(detail=True, methods=["get"])
    def print(self, request, pk=None):
        return render(request, "interviews/print.html", self._get_print_context(self.get_object()))

    @action(detail=True, methods=["get"])
    def pdf(self, request, pk=None):
        from weasyprint import HTML
        interview = self.get_object()
        ctx = self._get_print_context(interview)
        html = render_to_string("interviews/print.html", ctx)
        pdf = HTML(string=html).write_pdf()
        emp = interview.employee
        filename = f"{interview.get_type_display()}_{emp.last_name}_{emp.first_name}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["post"])
    def upload_document(self, request, pk=None):
        interview = self.get_object()
        file = request.FILES.get("document")
        if file:
            if interview.document:
                interview.document.delete()
            interview.document = file
            interview.save(update_fields=["document"])
        return Response(self.get_serializer(interview).data)

    @action(detail=True, methods=["post"])
    def remove_document(self, request, pk=None):
        interview = self.get_object()
        if interview.document:
            interview.document.delete()
            interview.document = None
            interview.save(update_fields=["document"])
        return Response({"status": "ok"})

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"error": "Aucun ID fourni"}, status=status.HTTP_400_BAD_REQUEST)
        qs = self.get_queryset().filter(pk__in=ids)
        count = qs.count()
        qs.delete()
        return Response({"deleted": count})


class InterviewTemplateViewSet(viewsets.ModelViewSet):
    queryset = InterviewTemplate.objects.all()
    serializer_class = InterviewTemplateSerializer
    permission_classes = [permissions.IsAuthenticated, InterviewPermission]

    @action(detail=False, methods=["post"])
    def import_csv(self, request):
        import csv
        import io

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)
        upload_error = validate_csv_upload(file)
        if upload_error:
            return Response({"error": upload_error}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(file.read().decode("utf-8-sig")))
        rows = list(reader)

        template_questions = {}
        seen_types = set()
        created = 0
        errors = []

        for row in rows:
            name = row.get("name", "").strip()
            if not name:
                continue
            ttype = row.get("type", "").strip()
            if ttype and ttype not in seen_types:
                seen_types.add(ttype)
            description = row.get("description", "").strip()
            section_id = row.get("section_id", "s1").strip()
            section_title = row.get("section_title", "").strip()
            question_id = row.get("question_id", "").strip()
            question_label = row.get("question_label", "").strip()
            question_type = row.get("question_type", "textarea").strip()

            if not section_title and not question_label:
                continue

            key = (name, ttype, description)
            if key not in template_questions:
                template_questions[key] = {}

            if section_id not in template_questions[key]:
                template_questions[key][section_id] = {
                    "id": section_id,
                    "title": section_title or "Section",
                    "questions": [],
                }

            if question_label and question_id:
                template_questions[key][section_id]["questions"].append({
                    "id": question_id,
                    "label": question_label,
                    "type": question_type if question_type in ("textarea", "rating", "yesno", "table") else "textarea",
                    "answer": "",
                })

        for (name, ttype, description), sections_dict in template_questions.items():
            sections = list(sections_dict.values())
            try:
                InterviewTemplate.objects.create(
                    name=name,
                    type=ttype,
                    description=description,
                    sections=sections,
                )
                created += 1
            except Exception as e:
                errors.append(f"{name}: {str(e)}")

        return Response({"created": created, "errors": errors})


class CampaignViewSet(viewsets.ModelViewSet):
    queryset = Campaign.objects.prefetch_related("interviews")
    serializer_class = CampaignSerializer
    permission_classes = [permissions.IsAuthenticated, InterviewPermission]

    def perform_create(self, serializer):
        serializer.save()

    def perform_update(self, serializer):
        old = self.get_object()
        new = serializer.save()
        if old.due_date != new.due_date:
            new.interviews.all().update(due_date=new.due_date)

    @action(detail=True, methods=["post"])
    def delete_all_interviews(self, request, pk=None):
        campaign = self.get_object()
        count = campaign.interviews.count()
        campaign.interviews.all().delete()
        return Response({"deleted": count})

    @action(detail=True, methods=["post"])
    def generate(self, request, pk=None):
        campaign = self.get_object()
        if not campaign.template:
            return Response({"error": "La campagne n'a pas de modèle"}, status=status.HTTP_400_BAD_REQUEST)

        template = campaign.template
        qs = User.objects.filter(is_active=True).exclude(role__in=("admin", "rh")).exclude(statut__in=("inactif", "sortie"))

        pf = campaign.population_filter or {}
        site = pf.get("site")
        service = pf.get("service")
        employee_ids = pf.get("employees")

        if employee_ids:
            qs = qs.filter(id__in=employee_ids)
        else:
            if site:
                qs = qs.filter(site_id=site)
            if service:
                qs = qs.filter(service_id=service)

        if not qs.exists():
            return Response(
                {"error": "Aucun collaborateur ne correspond aux critères de ce filtre"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        for user in qs:
            sections = list(template.sections)
            bilan_sections = build_previous_year_bilan_sections(user, template.type)
            if bilan_sections:
                sections = bilan_sections + sections
            sections = apply_default_sections(sections)
            _, was_created = Interview.objects.get_or_create(
                campaign=campaign,
                employee=user,
                type=template.type,
                defaults={
                    "template": template,
                    "template_snapshot": list(template.sections),
                    "content": {
                        "sections": sections,
                        "employee_snapshot": {
                            "position": user.position.name if user.position else None,
                            "service": user.service.name if user.service else None,
                            "site": user.site.name if user.site else None,
                            "coefficient": user.coefficient,
                        },
                    },
                    "due_date": campaign.due_date,
                    "manager": user.manager,
                },
            )
            if was_created:
                created += 1

        return Response({"created": created, "total": qs.count()})

    @action(detail=True, methods=["post"])
    def reassign_managers(self, request, pk=None):
        campaign = self.get_object()
        updated = 0
        for iv in campaign.interviews.select_related("employee").all():
            new_manager = iv.employee.manager
            if iv.manager_id != (new_manager.id if new_manager else None):
                iv.manager = new_manager
                iv.save(update_fields=["manager"])
                updated += 1
        return Response({"updated": updated})

    @action(detail=False, methods=["get"])
    def export_xlsx(self, request):
        campaign_ids = request.query_params.getlist("campaign_ids[]")
        six_years_ago = timezone.now() - timedelta(days=365.25 * 6)
        current_year = timezone.now().year
        years = list(range(current_year - 6, current_year + 1))

        qs = Interview.objects.filter(
            status__in=("completed", "signed"),
            created_at__gte=six_years_ago,
        ).select_related("employee", "employee__site")

        if campaign_ids:
            qs = qs.filter(campaign_id__in=[int(c) for c in campaign_ids])

        TYPE_LABELS = dict(Interview.Type.choices)
        emp_by_type = {}
        for iv in qs.iterator():
            emp_by_type.setdefault(iv.type, {}).setdefault(iv.employee.id, []).append(iv)

        wb = Workbook()
        ws = wb.active
        ws.title = "Synthèse"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        oui_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        non_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        headers = [
            "Type d'entretien", "ID Collaborateur", "Nom et Prénom", "Site",
            "Tous les entretiens (6 ans)",
            "Au moins une formation",
            "Au moins une évolution salaire",
            "Au moins une évolution pro",
        ]
        for y in years:
            headers.append(f"Entretien {y}")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_idx = 2
        for iv_type, emp_dict in emp_by_type.items():
            type_label = TYPE_LABELS.get(iv_type, iv_type)

            for emp_id, emp_ivs in emp_dict.items():
                emp = emp_ivs[0].employee
                flags = {"has_formation": False, "has_salary": False, "has_evolution": iv_type == "professional"}
                years_interviewed = {y: False for y in years}

                for iv in emp_ivs:
                    iv_year = iv.created_at.year
                    if iv_year in years_interviewed:
                        years_interviewed[iv_year] = True
                    snap = iv.content.get("employee_snapshot", {})
                    if snap.get("salaire_brut"):
                        flags["has_salary"] = True
                    if not flags["has_formation"]:
                        for section in iv.content.get("sections", []):
                            for q in section.get("questions", []):
                                qid = q.get("id", "")
                                if any(kw in qid for kw in ["formation", "cpf", "vae", "certif"]):
                                    if q.get("answer"):
                                        flags["has_formation"] = True
                                        break
                            if flags["has_formation"]:
                                break

                ws.cell(row=row_idx, column=1, value=type_label).border = thin_border
                ws.cell(row=row_idx, column=2, value=emp.id).border = thin_border
                ws.cell(row=row_idx, column=3, value=emp.get_full_name() or emp.email).border = thin_border
                ws.cell(row=row_idx, column=4, value=emp.site.name if emp.site else "").border = thin_border
                ws.cell(row=row_idx, column=5, value="").border = thin_border

                for col_idx, key in [(6, "has_formation"), (7, "has_salary"), (8, "has_evolution")]:
                    val = "Oui" if flags[key] else "Non"
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = oui_fill if flags[key] else non_fill

                for y in years:
                    col_idx = 9 + years.index(y)
                    val = "Oui" if years_interviewed[y] else ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if years_interviewed[y]:
                        cell.fill = oui_fill

                row_idx += 1

            row_idx += 1

        last_row = row_idx - 1
        if last_row > 1:
            col_letter = chr(64 + len(headers)) if len(headers) <= 26 else 'A'
            ws.auto_filter.ref = f"A1:{col_letter}{last_row}"
        ws.freeze_panes = "B2"

        for col_idx in range(1, len(headers) + 1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'
            ws.column_dimensions[col_letter].width = 24

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "synthese_entretiens.xlsx"
        if campaign_ids:
            campaigns = Campaign.objects.filter(pk__in=campaign_ids)
            names = "_".join(c.name.replace(" ", "_") for c in campaigns)
            if names:
                filename = f"synthese_{names}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _build_contents_workbook(self, interviews_qs, filename_base):
        interviews = interviews_qs.select_related(
            "employee", "employee__site", "employee__service", "employee__position", "manager"
        )

        TYPE_LABELS = dict(Interview.Type.choices)
        interviews_by_type = {}
        for iv in interviews:
            interviews_by_type.setdefault(iv.type, []).append(iv)

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        base_headers = ["ID", "Nom", "Prénom", "Email", "Site", "Service", "Poste", "Manager", "Statut entretien", "Date limite"]

        first_sheet = True
        for iv_type, type_interviews in interviews_by_type.items():
            sheet_name = TYPE_LABELS.get(iv_type, iv_type)[:31]
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            questions_in_type = []
            seen_qids = set()
            for iv in type_interviews:
                for section in iv.content.get("sections", []):
                    for q in section.get("questions", []):
                        qid = q.get("id", "")
                        if qid and qid not in seen_qids:
                            seen_qids.add(qid)
                            questions_in_type.append((qid, q.get("label", qid)))

            headers = base_headers + [label for _qid, label in questions_in_type]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, iv in enumerate(type_interviews, 2):
                emp = iv.employee
                row_data = [
                    emp.id, emp.last_name, emp.first_name, emp.email,
                    emp.site.name if emp.site else "",
                    emp.service.name if emp.service else "",
                    emp.position.name if emp.position else "",
                    (iv.manager.get_full_name() or iv.manager.email) if iv.manager else "",
                    iv.get_status_display(),
                    str(iv.due_date),
                ]
                answers = {}
                for section in iv.content.get("sections", []):
                    for q in section.get("questions", []):
                        qid = q.get("id", "")
                        answer = q.get("answer")
                        if qid:
                            if q.get("type") == "rating":
                                answers[qid] = str(answer) if answer is not None else ""
                            elif q.get("type") == "table" and isinstance(answer, list):
                                answers[qid] = "; ".join(
                                    " | ".join(str(c) for c in row) for row in answer if row
                                )
                            else:
                                answers[qid] = str(answer) if answer else ""

                for _qid, label in questions_in_type:
                    row_data.append(answers.get(_qid, ""))

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_cell_value(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            if type_interviews:
                last_col = len(headers)
                last_row = len(type_interviews) + 1
                col_letter = chr(64 + last_col) if last_col <= 26 else 'A'
                ws.auto_filter.ref = f"A1:{col_letter}{last_row}"
            ws.freeze_panes = "B2"

            for col_idx, col in enumerate(ws.columns, 1):
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{filename_base}.xlsx".replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=["get"])
    def export_contents_xlsx(self, request, pk=None):
        campaign = self.get_object()
        return self._build_contents_workbook(
            campaign.interviews.all(), f"contenus_{campaign.name}"
        )

    @action(detail=False, methods=["get"])
    def export_all_contents_xlsx(self, request):
        campaign_ids = request.query_params.getlist("campaign_ids[]")
        qs = Interview.objects.all()
        if campaign_ids:
            qs = qs.filter(campaign_id__in=[int(c) for c in campaign_ids])
        return self._build_contents_workbook(qs, "contenus_campagnes")
